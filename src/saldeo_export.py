"""
saldeo_export.py — Генерация Excel-файла импорта счетов-фактур для Saldeo Smart.

Использование:
    from saldeo_export import generate_saldeo_xlsx
    path, warnings = generate_saldeo_xlsx(df, "faktury.xlsx", start_inv_num=9)
"""

import calendar
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


# ── Реквизиты продавца ────────────────────────────────────────────────────────
# Личные данные продавца (имя, счёт, название услуги) НЕ хранятся в коде —
# пользователь вводит их один раз в мастере первого запуска, и они
# сохраняются в config.json. Здесь — только нейтральные значения по
# умолчанию на случай, если функция вызвана без явных параметров.

SELLER_NAME    = ""
SELLER_ACCOUNT = ""

SERVICE_NAME   = "Usługa"
VAT_RATE       = "zw."
VAT_BASIS      = "a113"   # Допустимые значения: a43, a113, a82, du, iz
PAYMENT_METHOD = "przelew"
UNIT           = "szt."
CURRENCY       = "PLN"
COUNTRY        = "PL"

# По умолчанию никто не исключён — пользователь снимает галочки сам
DEFAULT_EXCLUDE: set[str] = set()

# Порядок колонок шаблона Saldeo (50 штук) — важен для импорта
SALDEO_COLUMNS = [
    "Lp.",
    "Sufiks",
    "Własny opis faktury",
    "Metoda kasowa",
    "VAT marża",
    "Data wystawienia",
    "Data dostawy",
    "Termin płatności",
    "Wyślij do KSeF",
    "Wyślij do kontrahenta",
    "Nabywca (nazwa skrócona kontrahenta)",
    "Nazwa pełna kontrahenta",
    "Adres",
    "Kod pocztowy",
    "Miejscowość",
    "NIP",
    "REGON",
    "Kraj - skrót kraju w formacie ISO",
    "Adres e-mail",
    "Odbiorca (nazwa skrócona kontrahenta)",
    "Nazwa pełna (odbiorcy)",
    "Adres (odbiorcy)",
    "Kod pocztowy (odbiorcy)",
    "Miejscowość (odbiorcy)",
    "NIP (odbiorcy)",
    "REGON (odbiorcy)",
    "Kraj - skrót kraju w formacie ISO (odbiorcy)",
    "Adres e-mail (odbiorcy)",
    "Waluta",
    "Nazwa towaru",
    "PKWiU",
    "Ilość",
    "Jednostka",
    "Cena jedn. netto",
    "Cena jedn. brutto",
    "Stawka VAT",
    "Podstawa zastosowania stawki ZW",
    "Forma płatności",
    "Zapłacono",
    "Data wpłaty",
    "Konto bankowe",
    "Rodzaj środka transportu",
    "Data dopuszczenia",
    "Liczba rbh / Przebieg pojazdu",
    "Uwagi",
    "Wystawca faktury",
    "Grupa Towarowa",
    "Kategoria",
    "Opis",
    "Rejestr",
]


# ── Вспомогательные функции ───────────────────────────────────────────────────

def _last_day_of_month(d: date) -> date:
    """Последний день месяца для даты d."""
    last = calendar.monthrange(d.year, d.month)[1]
    return d.replace(day=last)


def _parse_date(s: str) -> date:
    """Парсит строку даты 'YYYY-MM-DD' или 'DD.MM.YYYY'."""
    s = s.strip()
    try:
        return date.fromisoformat(s)          # YYYY-MM-DD
    except ValueError:
        day, mon, yr = s.split(".")
        return date(int(yr), int(mon), int(day))


# ── Основная функция ──────────────────────────────────────────────────────────

def generate_saldeo_xlsx(
    df: pd.DataFrame,
    output_path: str,
    start_inv_num: int = 1,
    exclude_names: set | None = None,
    name_overrides: dict[str, str] | None = None,
    vat_basis: str | None = None,
    seller_name: str | None = None,
    seller_account: str | None = None,
    service_name: str | None = None,
) -> tuple[str, list[str]]:
    """
    Генерирует Excel-файл импорта фактур для Saldeo Smart.

    Parameters
    ----------
    df             : DataFrame с колонками date, name, amount, title
                     (результат parser.analyze())
    output_path    : путь для сохранения .xlsx
    start_inv_num  : номер первой фактуры (Lp. в файле)
    exclude_names  : имена клиентов для исключения; None → DEFAULT_EXCLUDE
    name_overrides : сопоставление {имя_из_выписки: каноническое_имя_из_Saldeo} —
                     применяется ПОСЛЕ исключения клиентов; позволяет привести
                     написание имени к тому, что уже есть в справочнике
                     контрагентов Saldeo, и тем самым избежать создания
                     дублирующей карточки контрагента при импорте.
                     None → подмена не выполняется
    vat_basis      : текст поля «Podstawa zastosowania stawki ZW»;
                     None → VAT_BASIS (константа)
    seller_name    : имя/название продавца («Wystawca faktury»);
                     None → SELLER_NAME (константа)
    seller_account : номер банковского счёта продавца;
                     None → SELLER_ACCOUNT (константа)
    service_name   : название услуги («Nazwa towaru»);
                     None → SERVICE_NAME (константа)

    Returns
    -------
    (output_path, warnings_list)
      warnings_list — список строк предупреждений (может быть пустым)
    """
    if exclude_names is None:
        exclude_names = DEFAULT_EXCLUDE
    if vat_basis is None:
        vat_basis = VAT_BASIS
    if seller_name is None:
        seller_name = SELLER_NAME
    if seller_account is None:
        seller_account = SELLER_ACCOUNT
    if service_name is None:
        service_name = SERVICE_NAME

    # Фильтруем исключённых клиентов
    df = df[~df["name"].isin(exclude_names)].copy()

    # Приводим написание имён к каноническому виду из справочника Saldeo
    # (применяется только к оставшимся, невыключенным клиентам — порядок важен:
    # пользователь исключает клиентов по исходным именам из выписки).
    if name_overrides:
        df["name"] = df["name"].replace(name_overrides)

    if df.empty:
        return output_path, [
            "Brak klientów do wygenerowania faktur (wszyscy wykluczeni lub brak danych)."
        ]

    # Разбираем даты в объекты date
    df["_date_obj"] = df["date"].apply(_parse_date)

    # Дата выставления = последний день месяца из данных
    issue_date = _last_day_of_month(df["_date_obj"].max())

    rows: list[dict] = []
    warnings_list: list[str] = []
    inv_num = start_inv_num

    # Клиенты в хронологическом порядке первого платежа
    first_dates = df.groupby("name")["_date_obj"].min()
    clients_ordered = first_dates.sort_values().index.tolist()

    for name in clients_ordered:
        grp = df[df["name"] == name].sort_values("_date_obj")
        amounts   = grp["amount"].round(2).tolist()
        first_dt  = grp["_date_obj"].min()
        last_dt   = grp["_date_obj"].max()

        # Data dostawy — одна дата (последний платёж).
        # Saldeo не принимает диапазон дат в этом поле.
        data_dostawy = last_dt.strftime("%d.%m.%Y")

        # Количество и цена: одинаковые суммы → кол-во × цена; разные → 1 × сумма
        unique_amounts = set(amounts)
        if len(unique_amounts) == 1:
            ilosc = len(amounts)
            cena  = amounts[0]
        else:
            ilosc = 1
            cena  = round(sum(amounts), 2)
            warnings_list.append(
                f"⚠  {name}: różne kwoty wpłat {amounts} — "
                f"zapisano jako 1 szt. × {cena:.2f} PLN (kwota łączna). "
                "Sprawdź i popraw ręcznie po imporcie."
            )

        total = round(ilosc * cena, 2)

        row = dict.fromkeys(SALDEO_COLUMNS, "")
        row["Lp."]                                  = inv_num
        # Адрес: берём из первой строки, где он есть; иначе "-"
        def _first_nonempty(series):
            for v in series:
                if v and str(v).strip():
                    return str(v).strip()
            return "-"

        street = _first_nonempty(grp.get("addr_street", []))
        postal = _first_nonempty(grp.get("addr_postal", []))
        city   = _first_nonempty(grp.get("addr_city",   []))

        row["Data wystawienia"]                     = issue_date.strftime("%d.%m.%Y")
        row["Data dostawy"]                         = data_dostawy
        row["Termin płatności"]                     = issue_date.strftime("%d.%m.%Y")
        row["Nabywca (nazwa skrócona kontrahenta)"] = name
        row["Nazwa pełna kontrahenta"]              = name
        row["Adres"]                                = street
        row["Kod pocztowy"]                         = postal
        row["Miejscowość"]                          = city
        row["Kraj - skrót kraju w formacie ISO"]    = COUNTRY
        row["Waluta"]                               = CURRENCY
        row["Nazwa towaru"]                         = service_name
        row["Ilość"]                                = ilosc
        row["Jednostka"]                            = UNIT
        row["Cena jedn. netto"]                     = cena
        # Cena jedn. brutto — оставляем пустой: Saldeo принимает либо netto, либо brutto
        row["Stawka VAT"]                           = VAT_RATE
        row["Podstawa zastosowania stawki ZW"]      = vat_basis
        row["Forma płatności"]                      = PAYMENT_METHOD
        row["Zapłacono"]                            = total
        row["Konto bankowe"]                        = seller_account
        row["Wystawca faktury"]                     = seller_name

        rows.append(row)
        inv_num += 1

    # ── Создаём Excel ─────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Faktury"

    # Заголовок
    ws.append(SALDEO_COLUMNS)
    hdr_font  = Font(bold=True)
    hdr_fill  = PatternFill("solid", fgColor="BDD7EE")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 36
    for cell in ws[1]:
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align

    # Данные
    data_align = Alignment(vertical="center")
    for row in rows:
        ws.append([row[c] for c in SALDEO_COLUMNS])
        row_idx = ws.max_row
        for cell in ws[row_idx]:
            cell.alignment = data_align

    # Автоширина (не шире 40 символов)
    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 40)

    # Закрепляем первую строку
    ws.freeze_panes = "A2"

    wb.save(output_path)
    return output_path, warnings_list
